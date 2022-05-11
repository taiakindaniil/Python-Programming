from math import sin, cos, ceil
from util import *
import matplotlib.pyplot as plt
import numpy as np
import openpyxl

def create_excel_file(x: list, y: list):
    wb = openpyxl.Workbook()
    sheet = wb.active

    sheet.cell(1, 1, "x")
    sheet.cell(1, 2, "y")

    for i, (coord_x, coord_y) in enumerate(zip(x, y)):
        sheet.cell(2+i, 1, coord_x)
        sheet.cell(2+i, 2, coord_y)

    wb.save("data.xlsx")

def default_f(x):
    # return cos(x) + 0.1*cos(x**5)
    # return cos(x**2) + x**2
    return sin(x) + 0.1*sin(x**5)

# forecasts new value with linear approximation
# returns prognosis x and y lists
def prognosis_data(step, x_data: list, y_data: list):
    x_data_copy = x_data.copy(); y_data_copy = y_data.copy()

    a, b = mnk([x_data[-2], x_data[-1]], [y_data[-2], y_data[-1]])

    new_x = x_data[-1] + step
    new_y = a*new_x + b

    x_data_copy.append(new_x)
    y_data_copy.append(new_y)

    return x_data_copy, y_data_copy

 # отклонение (сигма)
def std_deviation(orig_data: list, data: list):
    n = len(data)
    summ = 0
    for i in range(n):
        summ += (data[i] - orig_data[i])**2
    return ((1/n) * summ) ** 0.5

if __name__ == "__main__":
    print("Enter input data")
    start = int(input("start: "))
    stop  = int(input("stop: "))
    step  = float(input("step: "))

    x_data = list(np.arange(start, stop+step, step))
    y_data = [default_f(x) for x in x_data]

    create_excel_file(x_data, y_data)

    smooth_y_data = MAW(y_data)

    sig = std_deviation(y_data, smooth_y_data)
    smooth_y_data = MAW(y_data, k=max(2, ceil(sig)))

    # a, b = mnk(x_data, smooth_y_data)
    # print(a, b)

    new_x_data, new_y_data = prognosis_data(step, x_data, y_data)
    new_smth_x_data, new_smth_y_data = prognosis_data(step, x_data, smooth_y_data)

    print(new_y_data)
    print(new_smth_y_data)

    # show what we've done
    fig, axs = plt.subplots(1)
    axs.set_title("Data")
    axs.plot(x_data, y_data, color="#FE7A60", label="Initial")
    # prognosis point
    axs.scatter(new_x_data[-1], new_y_data[-1], color="#FE7A60")
    axs.plot(x_data, smooth_y_data, color="#77D5FF", label="Antialised")
    # prognosis point
    axs.scatter(new_x_data[-1], new_smth_y_data[-1], color="#77D5FF")
    axs.legend()

    plt.text(0.8, 1, f"σ = {sig}")

    fig.tight_layout()
    plt.show()